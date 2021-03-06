import requests
import json
import calendar
import urllib
from datetime import date
from _datetime import timedelta, datetime
import openpyxl
city = 'Nashville'
state ='TN'

wb = openpyxl.load_workbook('AC Run Log.xlsx')
ws = wb.get_sheet_by_name('Data Input')

def main():
    
    if date.today == 'Monday':
        weekendData = WeekendCalc()
        print(weekendData)
        print('in weekend clause')
        PlaceinXl(weekendData)
    else:
        weekDayData = WeekdayCalc()
        print(weekDayData)
        print('in weekday clause')
        PlaceinXl(weekDayData)
def DateCreator(date,deltaDate):
    
    workingDate = date - timedelta(deltaDate)
    yesterdayYear = '20' + workingDate.strftime('%y')
    yesterdayDay = workingDate.strftime('%d')
    yesterdayMonth = workingDate.strftime('%m')
    
    return(yesterdayDay,yesterdayMonth, yesterdayYear)

def CreatURL (city, state, year, month, day):
    
    combinedDate = year + month + day
    urlString = ('http://api.wunderground.com/api/4ed4d4b887e23593/history_{}/q/{}/{}.json'.format(combinedDate,state, city ))
    return urlString

def RequestData(url):
    if Checkconnectivity(url):
        jsonDatat = requests.get(url)
        print(jsonDatat.status_code)
        weatherData = jsonDatat.json()
        return weatherData
    else:
        return ''
    
def ParseEnginge (data):
    
    try:
        keyedDataExtraction = data['history']['dailysummary']
    except KeyError:
        return (0,0)
    
    for items in keyedDataExtraction:
        meantemp = items['meantempi']
        degreeDaysCooling =(items['coolingdegreedays'])
        return(meantemp, degreeDaysCooling)   

def WeekendCalc():
    
    createdDateSunday = DateCreator(date.today(), 1)
    createdDateSaturday = DateCreator(date.today(), 2)
    createUrlSunday = CreatURL(city, state, createdDateSunday[2], createdDateSunday[1], createdDateSunday[0])
    createUrlSaturday = CreatURL(city, state, createdDateSaturday[2], createdDateSaturday[1], createdDateSaturday[0])    
    sundayData = ParseEnginge(RequestData(createUrlSunday))
    satudayData = ParseEnginge(RequestData(createUrlSaturday))
    sundayMeanTemp = sundayData[0]
    sundayDegreeDays = sundayData[1]
    saturdayMeanTemp = satudayData[0]
    saturdayDegreeDays = satudayData[1]
    
    return (((sundayMeanTemp + saturdayMeanTemp) / 2 ),( sundayDegreeDays + saturdayDegreeDays))
    
def WeekdayCalc():
    createdDate = DateCreator(date.today(), 1)
    weekdayUrl = CreatURL(city, state, createdDate[2], createdDate[1], createdDate[0])
    weekdayData = ParseEnginge(RequestData(weekdayUrl))
    
    return weekdayData

def Checkconnectivity(reference):
    try:
        urllib.request.urlopen(reference, timeout=1)
        return True
    except urllib.request.URLError:
        return False

def PlaceinXl(wkData):
    PreviousDay = ws.max_row  #gets number of last row and increments by 1 to move to the next empty line
    cell1 = "N" + str(PreviousDay)
    cell2 = "O" + str(PreviousDay)
    cell3 = "A" + str(PreviousDay)
    ws[cell1] = int(wkData[0])
    ws[cell2] = int(wkData[1])
    day = str(ws[cell3].value)
    wb.save("AC Run Log.xlsx")
    print('complete')
    
main()
